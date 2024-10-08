import os

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_info_cell(worksheet, row, column, text):
    cell = worksheet.cell(row=row, column=column, value=text)

    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.border = thin_border

    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    column_letter = get_column_letter(column)
    max_text_length = max(len(line) for line in text.split("\n"))
    worksheet.column_dimensions[column_letter].width = max_text_length

    return cell


def create_excel_sheet(data_to_export, board_name, output_dir=None):
    data_export_df = pd.DataFrame(data_to_export)
    workbook = openpyxl.load_workbook("./csv/trello_template.xlsx")
    worksheet = workbook.active
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    current_list_value = None
    n_row = 1

    for _, card in data_export_df.iterrows():
        list_value = card["List"]

        if list_value != current_list_value:
            current_list_value = list_value
            n_row += 1
            for col in range(1, 7):
                cell = worksheet.cell(row=n_row, column=col)
                cell.fill = grey_fill

        n_row += 1

        for col in [1, 6]:
            cell = worksheet.cell(row=n_row, column=col)
            cell.fill = grey_fill

        worksheet.cell(row=n_row, column=2, value=card["List"])
        worksheet.cell(
            row=n_row, column=3, value=card["Name"]
        ).alignment = text_alignment
        worksheet.cell(
            row=n_row, column=4, value=card["Description"]
        ).alignment = text_alignment
        worksheet.cell(
            row=n_row, column=5, value=card["Labels"]
        ).alignment = text_alignment

    create_info_cell(
        worksheet,
        row=29,
        column=12,
        text="""
                    ADD YOUR HARDCODED ADDITIONAL INFORMATION HERE.
                    See excel.py line 74
                    """,
    )

    # Insert a closing grey row at the very end
    n_row += 1
    for col in range(1, 7):
        cell = worksheet.cell(row=n_row, column=col)
        cell.fill = grey_fill

    sanitized_board_name = "".join(
        c for c in board_name if c.isalnum() or c in (" ", "_")
    )
    output_dir = output_dir or "./csv"

    sanitized_filename = os.path.join(
        output_dir, f"{sanitized_board_name}_trello_template.xlsx"
    )
    workbook.save(sanitized_filename)
