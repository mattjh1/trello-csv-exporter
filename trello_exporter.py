### DEPRECATED ###
### DEPRECATED ###
### DEPRECATED ###
### DEPRECATED ###

import os
import sys
import time

import openpyxl
import pandas as pd
import requests
from dotenv import load_dotenv
from loguru import logger
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()


def create_info_cell(worksheet, row, column, text):
    cell = worksheet.cell(row=row, column=column, value=text)

    # Fill color
    cell.fill = PatternFill(
        start_color="FFE699", end_color="FFE699", fill_type="solid"
    )  # Yellow fill color

    # Add a border around the cell
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.border = thin_border

    # Alignment and wrap text
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Autosize the column to fit the widest text in the cell
    column_letter = get_column_letter(column)
    max_text_length = max(len(line) for line in text.split("\n"))
    worksheet.column_dimensions[column_letter].width = max_text_length

    return cell


def get_trello_boards(api_key, access_token):
    url = "https://api.trello.com/1/members/me/boards"

    params = {
        "key": api_key,
        "token": access_token,
    }

    response = requests.get(url, params=params)

    if response.status_code == 200:
        boards = response.json()
        return {board["name"]: board["id"] for board in boards}
    else:
        print(f"Failed to fetch Trello boards. Status code: {response.status_code}")
        return None


def select_trello_board(api_key, access_token):
    while True:
        boards = get_trello_boards(api_key, access_token)

        if boards:
            print("Available Trello boards:")
            for index, (board_name, board_id) in enumerate(boards.items(), start=1):
                print(f"{index}. {board_name}")

            selection = input("Enter the number of the board you want to export: ")
            print()

            try:
                selection = int(selection)
                if 1 <= selection <= len(boards):
                    selected_board = list(boards.values())[selection - 1]
                    return selected_board
                else:
                    print("Invalid selection. Please enter a valid number.")
                    time.sleep(2)
            except ValueError:
                print("Invalid input. Please enter a number.")
                time.sleep(2)
        else:
            print("No Trello boards found. Please check your credentials")
            sys.exit(1)


def get_trello_board_data(api_key, access_token, board_id):
    url = f"https://api.trello.com/1/boards/{board_id}"

    params = {
        "key": api_key,
        "token": access_token,
        "cards": "all",
        "lists": "all",
        "labels": "all",
    }

    response = requests.get(url, params=params)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"Failed to fetch board data. Status code: {response.status_code}")
        return None


def extract_card_data(board_data):
    data_to_export = []

    board_name = board_data["name"]

    for card in board_data["cards"]:
        list_name = None
        for list in board_data["lists"]:
            if list["id"] == card["idList"]:
                list_name = list["name"]
                break

        if list_name != "Maintenance":
            card_data = {
                "Name": card["name"],
                "Description": card["desc"],
                "List": list_name,
                "Labels": ", ".join([label["name"] for label in card["labels"]]),
            }
            data_to_export.append(card_data)

    return data_to_export, board_name


def create_excel_sheet(data_to_export, board_name):
    data_export_df = pd.DataFrame(data_to_export)
    data_export_df.sort_values(by="List", inplace=True)
    list_counts = data_export_df["List"].value_counts().reset_index()
    list_counts.columns = ["List", "Count"]
    sorted_df = data_export_df.merge(list_counts, on="List")
    sorted_df.sort_values(by=["Count", "List"], ascending=[False, True], inplace=True)

    workbook = openpyxl.load_workbook("./csv/trello_template.xlsx")
    worksheet = workbook.active

    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    current_list_value = None
    n_row = 1

    for _, card in sorted_df.iterrows():
        list_value = card["List"]

        if list_value != current_list_value:
            current_list_value = list_value
            n_row += 1
            for col in range(1, 7):
                cell = worksheet.cell(row=n_row, column=col)
                cell.fill = grey_fill

        n_row += 1

        for col in [1, 6]:
            cell = worksheet.cell(row=n_row, column=col)
            cell.fill = grey_fill

        worksheet.cell(row=n_row, column=2, value=card["List"])
        worksheet.cell(
            row=n_row, column=3, value=card["Name"]
        ).alignment = text_alignment
        worksheet.cell(
            row=n_row, column=4, value=card["Description"]
        ).alignment = text_alignment
        worksheet.cell(
            row=n_row, column=5, value=card["Labels"]
        ).alignment = text_alignment

    create_info_cell(
        worksheet,
        row=29,
        column=12,
        text="""
                    ADD YOUR HARDCODED ADDITIONAL INFORMATION HERE.
                    See trello_exporter.py line 161
                    """,
    )

    # Insert a closing grey row at the very end
    n_row += 1
    for col in range(1, 7):
        cell = worksheet.cell(row=n_row, column=col)
        cell.fill = grey_fill

    sanitized_board_name = "".join(
        c for c in board_name if c.isalnum() or c in (" ", "_")
    )
    sanitized_filename = f"./csv/{sanitized_board_name}_trello_template.xlsx"
    workbook.save(sanitized_filename)


if __name__ == "__main__":
    api_key = os.getenv("TRELLO_API_KEY")
    access_token = os.getenv("TRELLO_TOKEN")
    logger.warning(
        "\n\nThe script expects the following env variables to be present and valid:\n---------------------------------\n\tTRELLO_API_KEY\n\tTRELLO_TOKEN\n---------------------------------\n"
    )

    selected_board_id = select_trello_board(api_key, access_token)
    if selected_board_id:
        board_data = get_trello_board_data(api_key, access_token, selected_board_id)
        if board_data:
            logger.info("Starting board CSV exporter")
            card_data, board_name = extract_card_data(board_data)
            logger.info(f"Extracted {len(card_data)} cards")
            create_excel_sheet(card_data, board_name)
            logger.info("Excel file populated successfully")
            logger.info("Script completed")
